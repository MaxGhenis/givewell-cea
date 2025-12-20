#!/usr/bin/env python3
"""Debug benchmark values across different charity spreadsheets."""
import openpyxl

def find_benchmark(filepath, name):
    """Derive the benchmark value from the spreadsheet calculations."""
    print(f"\n=== {name} ===")
    wb = openpyxl.load_workbook(filepath, data_only=True)

    # Find Simple CEA sheet
    ws = None
    for sheet_name in wb.sheetnames:
        if "simple cea" in sheet_name.lower():
            ws = wb[sheet_name]
            break

    if ws is None:
        print("No Simple CEA sheet found")
        return

    # Find the row numbers for cost per death, xbenchmark, and moral weight
    cost_per_death_row = None
    initial_x_row = None
    moral_weight_row = None

    for row in range(1, 50):
        label = ws.cell(row=row, column=5).value or ws.cell(row=row, column=1).value
        if label:
            label_lower = str(label).lower()
            if "cost per" in label_lower and "death" in label_lower:
                cost_per_death_row = row
            if "initial cost-effectiveness" in label_lower and ("xcash" in label_lower or "xbenchmark" in label_lower or "terms of" in label_lower):
                initial_x_row = row
            if "moral" in label_lower and ("value" in label_lower or "weight" in label_lower):
                moral_weight_row = row

    print(f"  Found rows: cost_per_death={cost_per_death_row}, initial_x={initial_x_row}, moral_weight={moral_weight_row}")

    if not all([cost_per_death_row, initial_x_row, moral_weight_row]):
        print("  Could not find all required rows")
        return

    # Get moral weight from Overall column (usually 8)
    moral_weight = None
    for c in range(8, 10):
        mw = ws.cell(row=moral_weight_row, column=c).value
        if mw and isinstance(mw, (int, float)) and mw > 50:
            moral_weight = mw
            break

    if moral_weight is None:
        print("  Could not find moral weight")
        return

    # Calculate benchmark from first data column
    for col in range(9, 15):
        cost_per_death = ws.cell(row=cost_per_death_row, column=col).value
        initial_x = ws.cell(row=initial_x_row, column=col).value

        if cost_per_death and initial_x:
            if isinstance(cost_per_death, (int, float)) and isinstance(initial_x, (int, float)) and cost_per_death > 0 and initial_x > 0:
                # benchmark = moralWeight / (costPerDeath * xBenchmark)
                benchmark = moral_weight / (cost_per_death * initial_x)
                print(f"  Column {col}: cost_per_death={cost_per_death:.2f}, initial_x={initial_x:.4f}, moral_weight={moral_weight:.4f}")
                print(f"  Derived benchmark: {benchmark:.8f}")
                return benchmark

    print("  Could not derive benchmark")
    return None

# Check all spreadsheets
benchmarks = {}
for fname, name in [
    ("data/givewell-amf.xlsx", "AMF"),
    ("data/givewell-malaria-consortium.xlsx", "Malaria Consortium"),
    ("data/givewell-helen-keller.xlsx", "Helen Keller"),
    ("data/givewell-new-incentives.xlsx", "New Incentives"),
]:
    b = find_benchmark(fname, name)
    if b:
        benchmarks[name] = b

print("\n\n=== Summary ===")
for name, b in benchmarks.items():
    print(f"{name}: {b:.8f}")
