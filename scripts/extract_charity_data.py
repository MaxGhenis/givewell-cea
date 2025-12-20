#!/usr/bin/env python3
"""Extract charity data from GiveWell November 2025 CEA spreadsheets."""
import openpyxl

def examine_spreadsheet(filepath, name):
    print(f"\n\n{'='*60}")
    print(f"=== {name} ===")
    print('='*60)

    wb = openpyxl.load_workbook(filepath, data_only=True)

    print("\nSheets:")
    for sheet_name in wb.sheetnames:
        print(f"  - {sheet_name}")

    # Look for Simple CEA sheet
    for sheet_name in wb.sheetnames:
        if "simple cea" in sheet_name.lower():
            print(f"\n--- {sheet_name} ---")
            ws = wb[sheet_name]
            print(f"Dimensions: {ws.dimensions}")

            # Print first 45 rows
            for row in range(1, min(46, ws.max_row + 1)):
                vals = []
                for col in range(1, min(15, ws.max_column + 1)):
                    v = ws.cell(row=row, column=col).value
                    if v is not None:
                        s = str(v)[:30]
                        vals.append(f"{col}:{s}")
                if vals:
                    print(f"Row {row}: {vals}")
            break

examine_spreadsheet("data/givewell-helen-keller.xlsx", "Helen Keller")
print("\n\n")
examine_spreadsheet("data/givewell-new-incentives.xlsx", "New Incentives")
