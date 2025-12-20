#!/usr/bin/env python3
"""Extract test data from GiveWell spreadsheets for each charity."""
import openpyxl
import json

def get_cell_float(ws, row, col):
    """Get cell value as float, handling None and small numbers."""
    val = ws.cell(row=row, column=col).value
    if val is None:
        return None
    return float(val)

def extract_malaria_consortium():
    """Extract Malaria Consortium test data."""
    wb = openpyxl.load_workbook("data/givewell-malaria-consortium.xlsx", data_only=True)
    ws = wb["Simple CEA"]

    # Column mapping: 9=Burkina Faso, 10=Chad, 11=Côte d'Ivoire
    regions = {
        "Burkina Faso": 9,
        "Chad": 10,
        "Cote_dIvoire": 11,
    }

    data = {}
    for name, col in regions.items():
        data[name] = {
            "inputs": {
                "grantSize": 1_000_000,  # Normalized to $1M for comparison
                "costPerChildReached": get_cell_float(ws, 10, col),
                "malariaMortalityRate": get_cell_float(ws, 14, col),
                "proportionMortalityDuringSeason": get_cell_float(ws, 15, col),
                "smcEffect": get_cell_float(ws, 16, col),
                "moralWeightUnder5": get_cell_float(ws, 21, 8),  # Column 8 Overall
                "adjustmentOlderMortalities": get_cell_float(ws, 25, col),
                "adjustmentDevelopmental": get_cell_float(ws, 26, col),
                "adjustmentProgramBenefits": get_cell_float(ws, 34, col),
                "adjustmentGrantee": get_cell_float(ws, 35, col),
                "adjustmentLeverage": get_cell_float(ws, 36, col),
                "adjustmentFunging": get_cell_float(ws, 37, col),
            },
            "expected": {
                "initialXBenchmark": get_cell_float(ws, 22, col),
                "finalXBenchmark": get_cell_float(ws, 40, col),
            }
        }

    print("=== Malaria Consortium ===")
    print(json.dumps(data, indent=2))
    return data


def extract_helen_keller():
    """Extract Helen Keller test data."""
    wb = openpyxl.load_workbook("data/givewell-helen-keller.xlsx", data_only=True)
    ws = wb["Simple CEA"]

    # Column mapping: 9=Burkina Faso, 10=Cameroon, 12=DRC
    regions = {
        "Burkina_Faso": 9,
        "Cameroon": 10,
        "DRC": 12,
    }

    data = {}
    for name, col in regions.items():
        data[name] = {
            "inputs": {
                "grantSize": 1_000_000,
                "costPerPersonUnder5": get_cell_float(ws, 10, col),
                "proportionReachedCounterfactual": get_cell_float(ws, 12, col),
                "mortalityRateUnder5": get_cell_float(ws, 14, col),
                "vasEffect": get_cell_float(ws, 15, col),
                "moralWeightUnder5": get_cell_float(ws, 20, 8),  # Column 8 Overall
                "adjustmentDevelopmental": get_cell_float(ws, 24, col),
                "adjustmentProgramBenefits": get_cell_float(ws, 31, col),
                "adjustmentGrantee": get_cell_float(ws, 32, col),
                "adjustmentLeverage": get_cell_float(ws, 33, col),
                "adjustmentFunging": get_cell_float(ws, 34, col),
            },
            "expected": {
                "peopleUnder5Reached": get_cell_float(ws, 11, col),
                "additionalChildrenCovered": get_cell_float(ws, 13, col),
                "deathsAvertedUnder5": get_cell_float(ws, 16, col),
                "costPerDeathAverted": get_cell_float(ws, 19, col),
                "initialXBenchmark": get_cell_float(ws, 21, col),
                "finalXBenchmark": get_cell_float(ws, 37, col),
                "finalCostPerLifeSaved": get_cell_float(ws, 38, col),
            }
        }

    print("\n\n=== Helen Keller ===")
    print(json.dumps(data, indent=2))
    return data


def extract_new_incentives():
    """Extract New Incentives test data."""
    wb = openpyxl.load_workbook("data/givewell-new-incentives.xlsx", data_only=True)
    ws = wb["Simple CEA"]

    # Column mapping: 9=Bauchi, 10=Gombe, 11=Jigawa
    regions = {
        "Bauchi": 9,
        "Gombe": 10,
        "Jigawa": 11,
    }

    data = {}
    for name, col in regions.items():
        data[name] = {
            "inputs": {
                "grantSize": 1_000_000,
                "costPerChildReached": get_cell_float(ws, 10, 8),  # Overall
                "proportionReachedCounterfactual": get_cell_float(ws, 12, col),
                "probabilityDeathUnvaccinated": get_cell_float(ws, 14, col),
                "vaccineEffect": get_cell_float(ws, 15, col),
                "moralWeightUnder5": get_cell_float(ws, 20, 8),  # Overall
                "adjustmentOlderMortalities": get_cell_float(ws, 24, col),
                "adjustmentDevelopmental": get_cell_float(ws, 25, col),
                "adjustmentConsumption": get_cell_float(ws, 26, col),
                "adjustmentProgramBenefits": get_cell_float(ws, 35, col),
                "adjustmentGrantee": get_cell_float(ws, 36, 8),  # Overall
                "adjustmentLeverage": get_cell_float(ws, 37, col),
                "adjustmentFunging": get_cell_float(ws, 38, col),
            },
            "expected": {
                "childrenReached": get_cell_float(ws, 11, col),
                "additionalChildrenVaccinated": get_cell_float(ws, 13, col),
                "deathsAvertedUnder5": get_cell_float(ws, 16, col),
                "costPerDeathAverted": get_cell_float(ws, 19, col),
                "initialXBenchmark": get_cell_float(ws, 21, col),
                "finalXBenchmark": get_cell_float(ws, 41, col),
                "finalCostPerLifeSaved": get_cell_float(ws, 42, col),
            }
        }

    print("\n\n=== New Incentives ===")
    print(json.dumps(data, indent=2))
    return data


if __name__ == "__main__":
    mc = extract_malaria_consortium()
    hk = extract_helen_keller()
    ni = extract_new_incentives()

    # Save to JSON file
    all_data = {
        "malariaConsortium": mc,
        "helenKeller": hk,
        "newIncentives": ni,
    }

    with open("data/test_data.json", "w") as f:
        json.dump(all_data, f, indent=2)

    print("\n\nSaved to data/test_data.json")
