#!/usr/bin/env python3
"""Debug Malaria Consortium calculations."""
import openpyxl

wb = openpyxl.load_workbook("data/givewell-malaria-consortium.xlsx", data_only=True)
ws = wb["Simple CEA"]

# Burkina Faso column 9
col = 9

# Print key values from Simple CEA
print("=== Simple CEA Debug for Burkina Faso ===")
for row in range(1, 45):
    label = ws.cell(row=row, column=1).value or ws.cell(row=row, column=3).value or ws.cell(row=row, column=5).value
    if label:
        val = ws.cell(row=row, column=col).value
        val_overall = ws.cell(row=row, column=8).value
        print(f"Row {row}: {str(label)[:40]:40} | Col 9: {val} | Col 8: {val_overall}")

# Check the benchmark calculation
print("\n\n=== Calculating manually ===")
cost_per_child = ws.cell(row=10, column=col).value
mortality_rate = ws.cell(row=14, column=col).value
season_prop = ws.cell(row=15, column=col).value
smc_effect = ws.cell(row=16, column=col).value
moral_weight = ws.cell(row=21, column=8).value

print(f"Cost per child: {cost_per_child}")
print(f"Mortality rate: {mortality_rate}")
print(f"Season proportion: {season_prop}")
print(f"SMC effect: {smc_effect}")
print(f"Moral weight: {moral_weight}")

# Using grant size of 0.01 (from row 7)
grant_size = ws.cell(row=7, column=col).value
print(f"Grant size: {grant_size}")

children_reached = grant_size / cost_per_child
print(f"Children reached: {children_reached}")

deaths_averted = children_reached * mortality_rate * season_prop * smc_effect
print(f"Deaths averted: {deaths_averted}")

cost_per_death = grant_size / deaths_averted
print(f"Cost per death averted: {cost_per_death}")

# Calculate xbenchmark
value_generated = deaths_averted * moral_weight
print(f"Value generated: {value_generated}")

# Try different benchmark values
for benchmark in [0.003, 0.00333, 0.003355]:
    benchmark_value = grant_size * benchmark
    xbenchmark = value_generated / benchmark_value
    print(f"With benchmark {benchmark}: xBenchmark = {xbenchmark}")

# The expected is 14.43671251
print(f"\nExpected xBenchmark: 14.43671251")

# Let me also check the Inputs sheet for benchmark
print("\n\n=== Checking Inputs sheet ===")
ws_inputs = wb["Inputs"]
for row in range(1, 30):
    vals = []
    for col in range(1, 10):
        v = ws_inputs.cell(row=row, column=col).value
        if v:
            vals.append(f"{col}:{str(v)[:40]}")
    if vals:
        print(f"Row {row}: {vals}")
